# -*- coding: utf-8 -*-
"""Gera a folha de revisão clínica (Excel) do catálogo + propriedades + regras."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
cat = json.load(open(os.path.join(HERE, "catalogo.json"), encoding="utf-8"))

PETROL = "0F4C4A"; PSOFT = "E3ECEB"; YELLOW = "FBF0DF"; MUTED = "5D6C6B"
HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADY = Font(name="Arial", bold=True, color="7A5B12", size=11)
BODY = Font(name="Arial", size=10)
BODYB = Font(name="Arial", size=10, bold=True)
FILL = PatternFill("solid", fgColor=PETROL)
FILLY = PatternFill("solid", fgColor=YELLOW)
ALT = PatternFill("solid", fgColor="F3F6F6")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D7DEDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PROP_LABELS = {
 "qt":"prolonga o QT","cyp3a4i":"inibidor CYP3A4","cyp3a4s":"substrato CYP3A4",
 "cyp3a4ind":"indutor CYP3A4","estatina":"estatina","estatina3a4":"estatina (via CYP3A4)",
 "cyp2c19i":"inibidor CYP2C19","cyp2d6i":"inibidor CYP2D6","seroton":"serotoninérgico",
 "avnode":"deprime o nó AV","bbloq":"beta-bloqueante","colinergico":"colinérgico",
 "anticoag":"anticoagulante","avk":"antagonista da vit. K","antiagreg":"antiagregante",
 "clopidogrel":"clopidogrel","aine":"AINE","gi":"risco de úlcera/GI",
 "isrs_hemorr":"risco hemorrágico (ISRS)","inr_up":"aumenta o INR","kup":"retém potássio",
 "kdown":"baixa o potássio","ieca_ara":"bloqueio do SRAA","diur":"diurético","litio":"lítio",
 "digital":"digitálico","pgpi":"inibidor da P-gp","pgps":"substrato da P-gp",
 "snc":"depressor do SNC","opioide":"opioide","benzo":"benzodiazepina","anticol":"anticolinérgico",
 "nitrato":"nitrato","pde5":"inibidor da PDE5","alfa1blq":"alfa-bloqueante",
 "hipoglic":"risco de hipoglicemia","hipergli":"aumenta a glicemia","disglic":"disglicemia",
 "fibrato":"fibrato","tca":"tricíclico",
}
PROP_DEF = {
 "qt":"Prolonga o intervalo QT — risco de arritmia (torsades).",
 "cyp3a4i":"Inibidor do CYP3A4 — aumenta os níveis de substratos dessa via.",
 "cyp3a4s":"Substrato do CYP3A4 — os seus níveis sobem com inibidores.",
 "cyp3a4ind":"Indutor do CYP3A4 — reduz os níveis/eficácia de substratos.",
 "estatina":"Estatina (qualquer) — relevante para associação com fibratos.",
 "estatina3a4":"Estatina metabolizada pelo CYP3A4 (sinvastatina, atorvastatina) — risco de miopatia com inibidores.",
 "cyp2c19i":"Inibidor do CYP2C19.",
 "cyp2d6i":"Inibidor do CYP2D6.",
 "seroton":"Atividade serotoninérgica — risco de síndrome serotoninérgica em associação.",
 "avnode":"Deprime a condução no nó AV / bradicardizante.",
 "bbloq":"Beta-bloqueante.",
 "colinergico":"Efeito colinérgico (anticolinesterásico) — bradicardizante.",
 "anticoag":"Anticoagulante.",
 "avk":"Antagonista da vitamina K (varfarina, acenocumarol).",
 "antiagreg":"Antiagregante plaquetário.",
 "clopidogrel":"Clopidogrel — pró-fármaco ativado pelo CYP2C19.",
 "aine":"Anti-inflamatório não esteroide.",
 "gi":"Risco de lesão/úlcera gastrointestinal.",
 "isrs_hemorr":"Aumenta o risco hemorrágico (efeito serotoninérgico nas plaquetas).",
 "inr_up":"Potencia a varfarina (aumenta o INR).",
 "kup":"Retém potássio (risco de hipercaliemia).",
 "kdown":"Baixa o potássio (risco de hipocaliemia).",
 "ieca_ara":"Bloqueio do SRAA (IECA ou ARA).",
 "diur":"Diurético.",
 "litio":"Lítio — índice terapêutico estreito.",
 "digital":"Glicósido digitálico (digoxina).",
 "pgpi":"Inibidor da glicoproteína-P.",
 "pgps":"Substrato da glicoproteína-P (digoxina, DOAC, colchicina).",
 "snc":"Depressor do SNC / sedativo.",
 "opioide":"Opioide.",
 "benzo":"Benzodiazepina ou hipnótico análogo (Z-drug).",
 "anticol":"Carga anticolinérgica.",
 "nitrato":"Nitrato orgânico.",
 "pde5":"Inibidor da PDE5.",
 "alfa1blq":"Alfa-1-bloqueante.",
 "hipoglic":"Risco de hipoglicemia (antidiabético).",
 "hipergli":"Aumenta a glicemia (antagoniza o controlo glicémico).",
 "disglic":"Pode causar disglicemia (quinolona).",
 "fibrato":"Fibrato.",
 "tca":"Antidepressivo tricíclico.",
}

# regras do motor (mecanismo, dispara quando, severidade, mensagem, conduta)
RULES = [
 ("QT aditivo","2 substâncias que prolongam o QT","Maior","Prolongamento aditivo do QT — risco de torsades.","Evitar; vigiar ECG e eletrólitos (K, Mg)."),
 ("CYP3A4 + estatina","inibidor CYP3A4 + estatina (via CYP3A4)","Maior","Risco de miopatia/rabdomiólise.","Suspender ou trocar a estatina."),
 ("CYP3A4 + substrato","inibidor CYP3A4 + substrato CYP3A4","Moderada","Aumento da exposição do substrato.","Reduzir dose e vigiar toxicidade."),
 ("Indutor CYP3A4","indutor CYP3A4 + substrato/estatina 3A4/anticoagulante","Moderada","Reduz a exposição e a eficácia.","Vigiar eficácia; ajustar dose."),
 ("Serotoninérgica","2 agentes serotoninérgicos","Maior","Risco de síndrome serotoninérgica.","Evitar ou vigiar de perto."),
 ("Nó AV","2 agentes que deprimem o nó AV","Maior","Bradicardia e bloqueio AV.","Evitar; vigiar FC e ECG."),
 ("Colinérgico + bradicardizante","colinérgico + (deprime nó AV ou beta-bloqueante)","Moderada","Risco de bradicardia.","Vigiar FC."),
 ("Hemorragia (anticoagulante)","anticoagulante + (AINE ou antiagregante ou ISRS)","Maior","Risco hemorrágico elevado.","Rever necessidade; gastroproteção; vigiar."),
 ("Hemorragia (sem anticoag.)","antiagregante+AINE, ISRS+AINE, ou 2 antiagregantes","Moderada","Risco aumentado de hemorragia GI.","Ponderar gastroproteção; rever AINE."),
 ("INR (varfarina)","aumenta o INR + antagonista da vit. K","Moderada","Potencia a varfarina.","Vigiar INR durante e após."),
 ("Hipercaliemia","2 agentes que retêm potássio","Maior","Risco de hipercaliemia.","Vigiar potássio e função renal."),
 ("Hipercaliemia (AINE)","retém potássio + AINE","Moderada","Agrava a hipercaliemia e o risco renal.","Vigiar potássio e creatinina."),
 ("Hipocaliemia + digital","baixa o potássio + digitálico","Moderada","Potencia a toxicidade digitálica.","Vigiar potássio e digoxina."),
 ("Hipocaliemia + QT","baixa o potássio + prolonga o QT","Moderada","Potencia o prolongamento do QT.","Corrigir e vigiar o potássio."),
 ("Lítio","lítio + (AINE ou diurético ou bloqueio do SRAA)","Maior","Aumento dos níveis de lítio.","Vigiar litemia e função renal."),
 ("P-gp","inibidor da P-gp + substrato da P-gp","Moderada","Aumento dos níveis do substrato.","Reduzir dose e vigiar."),
 ("Opioide + benzodiazepina","opioide + benzodiazepina","Maior","Depressão do SNC e respiratória.","Evitar; doses mínimas e vigilância."),
 ("Depressão do SNC","2 depressores do SNC","Moderada","Sedação aditiva.","Cautela, sobretudo no idoso (quedas)."),
 ("Anticolinérgica","2 agentes anticolinérgicos","Moderada","Carga anticolinérgica aditiva.","Cautela no idoso."),
 ("Nitrato + PDE5","nitrato + inibidor da PDE5","Maior","Hipotensão grave.","Contraindicado — não associar."),
 ("Alfa-bloqueante + PDE5","alfa-bloqueante + inibidor da PDE5","Moderada","Hipotensão aditiva.","Separar tomas; iniciar dose baixa."),
 ("Beta-bloq. + hipoglicemia","beta-bloqueante + risco de hipoglicemia","Moderada","Mascara sintomas de hipoglicemia.","Alertar; reforçar auto-vigilância."),
 ("Quinolona + glicemia","disglicemia + risco de hipoglicemia","Moderada","Disglicemia.","Reforçar vigilância da glicemia."),
 ("Fibrato + estatina","fibrato + estatina","Moderada","Risco de miopatia.","Vigiar sintomas musculares e CK."),
 ("Corticoide + antidiabético","aumenta a glicemia + risco de hipoglicemia","Moderada","Antagoniza o controlo glicémico.","Reforçar vigilância; ajustar."),
 ("CYP2C19 + clopidogrel","inibidor CYP2C19 + clopidogrel","Moderada","Reduz a ativação do clopidogrel.","Preferir pantoprazol."),
 ("CYP2D6 + tricíclico","inibidor CYP2D6 + tricíclico","Moderada","Aumenta os níveis do tricíclico.","Reduzir dose e vigiar."),
 ("Triplo whammy","AINE + IECA/ARA + diurético (na lista)","Maior","Risco de lesão renal aguda.","Evitar o AINE; vigiar função renal."),
]

wb = Workbook()

def style_header(ws, ncols, review_from=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        if review_from and c >= review_from:
            cell.font = HEADY; cell.fill = FILLY
        else:
            cell.font = HEAD; cell.fill = FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

def body(ws, nrows, ncols):
    for r in range(2, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY; cell.alignment = WRAP; cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT

# ---- Instruções ----
ws0 = wb.active; ws0.title = "Instruções"
ws0.sheet_view.showGridLines = False
ws0["A1"] = "Revista — folha de revisão clínica"; ws0["A1"].font = Font(name="Arial", bold=True, size=16, color=PETROL)
notes = [
 "", "O que é: as substâncias do catálogo (base ATC) foram etiquetadas com PROPRIEDADES de mecanismo",
 "(ex.: prolonga o QT, inibidor do CYP3A4, serotoninérgico). O motor gera as interações a partir do",
 "cruzamento destas propriedades — por isso a exatidão das etiquetas é o que determina a qualidade.",
 "",
 "Como rever:",
 "  • Separador «Propriedades» — o significado de cada etiqueta. Confirma as definições.",
 "  • Separador «Substâncias» — 1 linha por fármaco, com as propriedades atribuídas. Nas colunas a",
 "    amarelo, marca se estão corretas e indica o que acrescentar/remover.",
 "  • Separador «Regras» — os ~27 mecanismos que disparam as interações. Marca se concordas.",
 "",
 "Preenche só as colunas a amarelo. Depois devolve o ficheiro e eu aplico as correções de uma vez.",
 "",
 "Aviso: conteúdo ilustrativo, de farmacologia clássica. Não se destina a uso clínico.",
]
for i, t in enumerate(notes, start=2):
    ws0.cell(row=i, column=1, value=t).font = Font(name="Arial", size=11,
        bold=t.endswith(":") or t.startswith("Como"))
ws0.column_dimensions["A"].width = 110

# ---- Propriedades ----
ws1 = wb.create_sheet("Propriedades")
ws1.append(["Chave", "Etiqueta", "Definição", "Manter? (S/N)", "Correção / notas"])
for k in sorted(PROP_LABELS):
    ws1.append([k, PROP_LABELS[k], PROP_DEF.get(k, ""), "", ""])
style_header(ws1, 5, review_from=4); body(ws1, ws1.max_row, 5)
for col, w in zip("ABCDE", [14, 24, 60, 14, 40]):
    ws1.column_dimensions[col].width = w

# ---- Substâncias ----
ws2 = wb.create_sheet("Substâncias")
ws2.append(["Grupo ATC", "Classe", "ATC classe", "Substância (DCI)", "ATC-5",
            "Propriedades atribuídas", "Propriedades OK? (S/N)", "Acrescentar", "Remover", "Notas"])
for c in cat["classes"]:
    for s in c["substancias"]:
        props = ", ".join(PROP_LABELS.get(p, p) for p in s["props"]) or "—"
        ws2.append([c["grupo_nome"], c["classe"], c["atc_classe"], s["dci"], s["atc"],
                    props, "", "", "", ""])
style_header(ws2, 10, review_from=7); body(ws2, ws2.max_row, 10)
for col, w in zip("ABCDEFGHIJ", [26, 34, 11, 24, 10, 40, 16, 22, 22, 30]):
    ws2.column_dimensions[col].width = w

# ---- Regras ----
ws3 = wb.create_sheet("Regras")
ws3.append(["#", "Mecanismo", "Dispara quando", "Severidade", "Mensagem", "Conduta",
            "Concorda? (S/N)", "Notas"])
for i, (nome, quando, sev, msg, cond) in enumerate(RULES, start=1):
    ws3.append([i, nome, quando, sev, msg, cond, "", ""])
style_header(ws3, 8, review_from=7); body(ws3, ws3.max_row, 8)
for col, w in zip("ABCDEFGH", [5, 28, 42, 12, 46, 40, 16, 30]):
    ws3.column_dimensions[col].width = w
# realçar severidade
sev_fill = {"Maior": "FAE8E5", "Moderada": "FBF0DF"}
for r in range(2, ws3.max_row + 1):
    v = ws3.cell(row=r, column=4).value
    if v in sev_fill:
        ws3.cell(row=r, column=4).fill = PatternFill("solid", fgColor=sev_fill[v])
        ws3.cell(row=r, column=4).font = BODYB

out = os.path.join(HERE, "revisao_clinica.xlsx")
wb.save(out)
print("gravado:", out)
print("substâncias:", ws2.max_row - 1, "| propriedades:", ws1.max_row - 1, "| regras:", ws3.max_row - 1)
