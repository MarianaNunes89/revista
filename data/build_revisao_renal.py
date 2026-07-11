# -*- coding: utf-8 -*-
"""Gera revisao_renal.xlsx — folha de revisão clínica dos alertas de ajuste à função renal."""
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HTML = "../index.html"
CAT  = "catalogo.json"

# extrair o dicionário RENAL do index.html
src = open(HTML, encoding="utf-8").read()
block = re.search(r"const RENAL = \{(.*?)\n\};", src, re.S).group(1)
RENAL = dict(re.findall(r'([A-Z0-9]{5,7})\s*:\s*"((?:[^"\\]|\\.)*)"', block))

cat = json.load(open(CAT, encoding="utf-8"))
info = {}
for c in cat["classes"]:
    for s in c["substancias"]:
        info[s["atc"]] = (s["dci"], c["classe"], c["grupo_atc"])

PRIOR = {  # risco se o limiar estiver errado
 "B01AE07":"Alto","A10BA02":"Alto","C01AA05":"Alto","M04AC01":"Alto","L01BA01":"Alto",
 "N05AN01":"Alto","B01AF02":"Alto","B01AF01":"Alto","B01AF03":"Alto","B01AB05":"Alto",
 "N02AA01":"Alto","J01XE01":"Alto","A10BB01":"Alto","M04AA01":"Alto","L04AX01":"Alto",
 "L01BB02":"Alto","J01EE01":"Alto",
 # reclassificados (Science): 100 % renais + depressão respiratória com opioides
 "N03AX12":"Alto","N03AX16":"Alto",
}

wb = Workbook(); wb.remove(wb.active)
HDR  = PatternFill("solid", fgColor="0F4C4A")
YEL  = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
BD   = Border(*[Side(style="thin", color="D9D9D9")]*4)
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)

# --- Instruções ---
ws = wb.create_sheet("Instruções")
ws.column_dimensions["A"].width = 118
linhas = [
 ("Revista — revisão clínica dos alertas de AJUSTE À FUNÇÃO RENAL (TFG)", True),
 ("", False),
 ("O que é isto: a ferramenta passou a mostrar um selo «TFG» nos fármacos que exigem ajuste à função renal.", False),
 ("Cada selo tem um texto com um limiar e uma conduta concretos. É a primeira vez que a ferramenta dá números de dose.", False),
 ("", False),
 ("O que preciso de si: validar cada linha da folha «Alertas renais». Preencha as colunas AMARELAS.", False),
 ("", False),
 ("  Confirma?      — Sim / Corrigir / Remover (alerta desnecessário, só gera ruído)", False),
 ("  Limiar correto — o valor de TFG está certo? Se não, qual?", False),
 ("  Conduta correta— reduzir para que dose / evitar / apenas vigiar?", False),
 ("  Fonte          — RCM, Stockley's, KDIGO, Renal Drug Handbook, etc.", False),
 ("  Nível          — consenso / provável / incerto", False),
 ("", False),
 ("Também importa: FALTA algum fármaco do catálogo que exija ajuste renal e não esteja aqui?", False),
 ("Anote-os na folha «Em falta».", False),
 ("", False),
 ("Prioridade: comece pelas linhas marcadas «Alto» na coluna Risco — são aquelas em que um limiar errado causa dano.", False),
 ("", False),
 ("Nota: a linagliptina foi deliberadamente EXCLUÍDA (excreção biliar, sem ajuste renal). Confirma?", False),
 ("", False),
 ("Conteúdo ilustrativo, não destinado a uso clínico.", False),
]
for i,(t,b) in enumerate(linhas, start=1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = Font(name="Arial", size=11, bold=b); c.alignment = Alignment(wrap_text=True, vertical="top")

# --- Alertas renais ---
ws = wb.create_sheet("Alertas renais")
cols = ["ATC5","Substância","Classe","Grupo","Risco se errado","Texto do alerta atual",
        "Confirma? (Sim/Corrigir/Remover)","Limiar correto","Conduta correta","Fonte","Nível"]
widths = [10,30,28,8,14,68,22,26,40,26,14]
for j,(h,w) in enumerate(zip(cols,widths), start=1):
    c = ws.cell(row=1, column=j, value=h); c.fill = HDR; c.font = WHITE_BOLD
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A2"

rows = sorted(RENAL.items(), key=lambda kv: (PRIOR.get(kv[0],"z") != "Alto", info.get(kv[0],("",))[0]))
r = 2
for atc, txt in rows:
    dci, cls, grp = info.get(atc, ("?","?","?"))
    vals = [atc, dci, cls, grp, PRIOR.get(atc,"Médio"), txt.replace('\\"','"'), "", "", "", "", ""]
    for j,v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = BOLD if j==5 and vals[4]=="Alto" else BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BD
        if j >= 7: c.fill = YEL
        elif vals[4]=="Alto" and j==5: c.fill = GREY
    ws.row_dimensions[r].height = 42
    r += 1

# --- Em falta ---
ws = wb.create_sheet("Em falta")
for j,(h,w) in enumerate(zip(["Substância (do catálogo)","Limiar de TFG","Conduta","Fonte","Nível"],[30,24,46,26,14]), start=1):
    c = ws.cell(row=1, column=j, value=h); c.fill = HDR; c.font = WHITE_BOLD
    c.alignment = Alignment(wrap_text=True, horizontal="center"); ws.column_dimensions[get_column_letter(j)].width = w
ex = ["(exemplo) Ranitidina","< 50","Reduzir a dose em 50 %","RCM","consenso"]
for j,v in enumerate(ex, start=1):
    c = ws.cell(row=2, column=j, value=v); c.font = Font(name="Arial", size=10, italic=True, color="888888"); c.fill = YEL

wb.save("revisao_renal.xlsx")
print("alertas:", len(RENAL), "| alto risco:", sum(1 for a in RENAL if PRIOR.get(a)=="Alto"))
